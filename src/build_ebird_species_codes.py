"""
Build a local eBird species-code mapping from the official taxonomy workbook.

The input workbook is the official "all common names" file linked from the
eBird support article "Bird Names in eBird". We only extract the fields the
review UI needs for reference-media lookup:

- species_code
- primary_com_name
- sci_name
- category

The resulting JSON is keyed by scientific name so the review workflow can map
its canonical labels to eBird species codes without runtime scraping.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path


def build_species_code_map(workbook_path: Path) -> dict[str, dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only in ad hoc usage
        raise SystemExit(
            "openpyxl is required to build the eBird species-code map. "
            "Install it with `pip install openpyxl`."
        ) from exc

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["full_sparse"]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = {name: idx for idx, name in enumerate(header)}

    mapping: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sci_name = row[columns["sci_name"]]
        species_code = row[columns["species_code"]]
        primary_com_name = row[columns["primary_com_name"]]
        category = row[columns["category"]]
        if not sci_name or not species_code:
            continue
        mapping[str(sci_name)] = {
            "species_code": str(species_code),
            "primary_com_name": str(primary_com_name or sci_name),
            "category": str(category or ""),
        }

    return mapping


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the official eBird taxonomy workbook (.xlsx).")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).parent.parent / "data" / "ebird_species_codes.json",
        help="Destination JSON file.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    mapping = build_species_code_map(args.workbook)
    payload = {
        "__meta__": {
            "source": args.workbook.name,
            "record_count": len(mapping),
        },
        **mapping,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, sort_keys=True)
        f.write("\n")


if __name__ == "__main__":
    main()
